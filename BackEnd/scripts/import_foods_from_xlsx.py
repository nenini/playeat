import argparse
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("openpyxl is required. Install it with: pip install openpyxl") from exc


COLUMN_ALIASES = {
    "external_food_code": ["식품코드", "제품코드"],
    "name": ["식품명", "제품명"],
    "brand": ["업체명", "제조사명"],
    "category": ["식품대분류명", "제품대분류명"],
    "nutrition_basis": ["영양성분함량기준량", "영양성분함량기준"],
    "serving_size": ["식품중량", "제품중량"],
    "calories": ["에너지(kcal)", "에너지"],
    "protein_g": ["단백질(g)", "단백질"],
    "carbs_g": ["탄수화물(g)", "탄수화물"],
    "fat_g": ["지방(g)", "지방"],
    "sugar_g": ["당류(g)", "당류"],
    "sodium_mg": ["나트륨(mg)", "나트륨"],
    "fiber_g": ["식이섬유(g)", "식이섬유"],
    "iron_mg": ["철(mg)", "철"],
    "phosphorus_mg": ["인(mg)", "인"],
    "potassium_mg": ["칼륨(mg)", "칼륨"],
    "vitamin_a_ug_rae": ["비타민A(㎍ RAE)", "비타민A"],
    "beta_carotene_ug": ["베타카로틴(㎍)", "베타카로틴"],
    "retinol_ug": ["레티놀(㎍)", "레티놀"],
    "source": ["출처명", "데이터출처명"],
}

NUTRIENT_COLUMNS = [
    "calories",
    "protein_g",
    "carbs_g",
    "fat_g",
    "sugar_g",
    "sodium_mg",
    "fiber_g",
    "iron_mg",
    "phosphorus_mg",
    "potassium_mg",
    "vitamin_a_ug_rae",
    "beta_carotene_ug",
    "retinol_ug",
]

NUMERIC_SQL_COLUMNS = set(NUTRIENT_COLUMNS) | {
    "nutrition_basis_amount",
    "serving_amount",
    "gram_per_piece",
}

INSERT_COLUMNS = [
    "external_food_code",
    "name",
    "brand",
    "category",
    "nutrition_basis_amount",
    "nutrition_basis_unit",
    "serving_amount",
    "serving_unit",
    "gram_per_piece",
    "calories",
    "protein_g",
    "carbs_g",
    "fat_g",
    "sugar_g",
    "sodium_mg",
    "fiber_g",
    "iron_mg",
    "phosphorus_mg",
    "potassium_mg",
    "vitamin_a_ug_rae",
    "beta_carotene_ug",
    "retinol_ug",
    "source",
]


def normalize_header(value):
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip())


def find_columns(header):
    normalized = {normalize_header(value): index for index, value in enumerate(header)}
    columns = {}
    for target, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            index = normalized.get(normalize_header(alias))
            if index is not None:
                columns[target] = index
                break
    missing = [name for name in ("external_food_code", "name") if name not in columns]
    if missing:
        raise SystemExit(f"Required columns are missing: {', '.join(missing)}")
    return columns


def cell(row, columns, name):
    index = columns.get(name)
    if index is None:
        return None
    value = row[index]
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def decimal_or_zero(value):
    if value is None:
        return Decimal("0")
    cleaned = str(value).replace(",", "").strip()
    if cleaned in {"", "-", "N/A"}:
        return Decimal("0")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return Decimal("0")


def parse_amount_unit(raw):
    if not raw:
        return None, None
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*([a-zA-Z가-힣]+)", str(raw))
    if not match:
        return None, None
    amount = Decimal(match.group(1))
    unit = normalize_unit(match.group(2))
    return amount, unit


def normalize_unit(raw_unit):
    unit = raw_unit.strip().lower()
    if unit in {"그램", "g"}:
        return "g"
    if unit in {"밀리리터", "ml"}:
        return "ml"
    return unit


def gram_per_piece_from_name(name, serving_amount, serving_unit):
    if serving_amount is None or serving_unit != "g":
        return None
    match = re.search(r"(\d+)\s*개입", name)
    if match:
        return serving_amount / Decimal(match.group(1))
    if "낱개" in name or "1개입" in name:
        return serving_amount
    return None


def sql_string(value):
    if value is None:
        return "NULL"
    cleaned = str(value).strip()
    if cleaned in {"", "해당없음"}:
        return "NULL"
    return "'" + cleaned.replace("\\", "\\\\").replace("'", "''") + "'"


def sql_number(value):
    if value is None:
        return "NULL"
    return str(value)


def food_values(row, columns):
    external_food_code = cell(row, columns, "external_food_code")
    name = cell(row, columns, "name")
    if not external_food_code or not name:
        return None

    basis_amount, basis_unit = parse_amount_unit(cell(row, columns, "nutrition_basis"))
    serving_amount, serving_unit = parse_amount_unit(cell(row, columns, "serving_size"))

    basis_amount = basis_amount or Decimal("100")
    basis_unit = basis_unit or "g"
    gram_per_piece = gram_per_piece_from_name(name, serving_amount, serving_unit)

    values = {
        "external_food_code": external_food_code,
        "name": name,
        "brand": cell(row, columns, "brand"),
        "category": cell(row, columns, "category"),
        "nutrition_basis_amount": basis_amount,
        "nutrition_basis_unit": basis_unit,
        "serving_amount": serving_amount,
        "serving_unit": serving_unit,
        "gram_per_piece": gram_per_piece,
        "source": cell(row, columns, "source"),
    }
    for column in NUTRIENT_COLUMNS:
        values[column] = decimal_or_zero(cell(row, columns, column))
    return values


def insert_statement(values):
    sql_values = []
    for column in INSERT_COLUMNS:
        value = values[column]
        if column in NUMERIC_SQL_COLUMNS:
            sql_values.append(sql_number(value))
        else:
            sql_values.append(sql_string(value))

    updates = ", ".join(
        f"{column} = VALUES({column})"
        for column in INSERT_COLUMNS
        if column != "external_food_code"
    )
    return (
        f"INSERT INTO foods ({', '.join(INSERT_COLUMNS)})\n"
        f"VALUES ({', '.join(sql_values)})\n"
        f"ON DUPLICATE KEY UPDATE {updates};"
    )


def main():
    parser = argparse.ArgumentParser(description="Convert food xlsx rows into MySQL seed SQL.")
    parser.add_argument("xlsx_path")
    parser.add_argument("--sheet")
    parser.add_argument("--output", default="BackEnd/scripts/generated/foods-seed.sql")
    args = parser.parse_args()

    workbook = load_workbook(args.xlsx_path, read_only=True, data_only=True)
    sheet = workbook[args.sheet] if args.sheet else workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    columns = find_columns(header)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    written = 0
    with output_path.open("w", encoding="utf-8", newline="\n") as output:
        output.write("-- Generated by BackEnd/scripts/import_foods_from_xlsx.py\n")
        for row in rows:
            values = food_values(row, columns)
            if values is None:
                continue
            output.write(insert_statement(values))
            output.write("\n")
            written += 1

    print(f"Wrote {written} food rows to {output_path}")


if __name__ == "__main__":
    main()
